import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import re
import difflib
from tkinter import *
import tkinterweb as th3
import tkinter.ttk
from tkinter.filedialog import *
from tkinter.messagebox import *
import threading
import webbrowser

'''#   登录界面
sub = 1


def form():  # 第1个窗体：登录窗体

    def ok():
        global sub
        if en1.get() == '123':
            root1.destroy()  # 关闭登录窗体
            sub = 0
        else:
            showwarning("警告：", "密码错!")

    root1 = Tk()
    root1.title('登录窗口')
    root1.geometry('300x150+600+200')
    la0 = Label(root1, text='请输入密码：')
    la0.pack()
    en1 = Entry(root1)
    en1.pack()
    but1 = Button(root1, text=" 确 定 ", command=ok)  # 判断密码是否正确
    but1.pack(pady=5)
    but2 = Button(root1, text=" 退 出 ", command=root1.destroy)  # 关闭登录窗体
    but2.pack(pady=5)

    root1.mainloop()  # 一直在等待接受窗体1事件,不会进入第2个窗体


form()
#   无限等待，直到密码输入正确跳转下一界面
while sub:
    pass'''

root = Tk()
root.title("text compare")
result_text: Text = Text(root)
filename1 = StringVar()
filename2 = StringVar()
filedir1 = StringVar()
filedir2 = StringVar()
save_path = StringVar()
mode_name = StringVar()
mode_name.set("操作模式")
project_name = '项目'
program_name = '类型'
signal = 0
# 默认查找的文件位置
default = None


class ProgramTemplate:
    def __init__(self, comments, logic_word, head_note, ex_note):
        self.comments = comments
        self.logic_word = logic_word
        self.head_note = head_note
        self.ex_note = ex_note

# 筛选出程序中的逻辑段
    def logic_segment(self):
        # sentences = re.split('\n', self.comments)  # 使用换行分割文本为句子列表
        logic_results = []
        # 找到开始标志在列表中的位置，并从此开始再筛选需要的逻辑段
        for index, item in enumerate(self.comments):
            if 'NOP\n' in item or '/MN\n' in item:
                for sentence in self.comments[index:]:
                    if re.match(self.logic_word, sentence):
                        logic_results.append(sentence)
                break

        return logic_results

# 筛选出程序抬头的备注部分
    def remarks(self):
        # sentences = re.split(r'\n', self.comments)  # 使用换行分割文本为句子列表
        ex_results = []
        for sentence in self.comments:
            if re.match(self.head_note, sentence):  # 匹配开头为字母或数字的句子
                ex_results.append(sentence)
        return ex_results

# 筛选出程序中特殊备注的部分
    def ex_part(self):
        # sentences = re.split(r'\n', self.comments)  # 使用换行分割文本为句子列表
        remarks_results = []

        for sentence in self.comments:
            if re.match(self.ex_note, sentence):  # 匹配开头为字母或数字的句子
                remarks_results.append(sentence)

        return remarks_results


# 为不同机器人文件类型选择相应的筛选关键字
def keyword(extension):
    keys = []
    # 发那科
    if extension == '.LS' or extension == '.ls':
        # 主要对比结果为第一个：包括程序抬头备注和主要逻辑段
        keys = [r'\s*\d+:\s*([$a-zA-Z0-9*]{2,}|[!/\-;]+\s*\*+\s*[a-zA-Z0-9]*|[!;]+\s*[a-zA-Z0-9]+\s*:)',
                r'\s*/*[a-zA-Z]+', r'\s*\d+:\s*[!/\-;]+\s*[a-zA-Z0-9]+']
    # 安川
    elif extension == '.JBI' or extension == '.jbi':
        keys = [r'\s*[$*a-zA-Z0-9*]{2,}\s*\d*', r'/+[a-zA-Z0-9]+', r'\'\s*[a-zA-Z0-9]+']
    return keys


# 选择输入文件路径文件
def input_file(file):
    filepath = askopenfilename()  # 选择打开什么文件，返回文件名
    if filepath.strip() != '':
        file.set(filepath)  # 设置变量filename1的值


# 选择输入文件夹路径
def input_dir(dir_path):
    file_dir_path = askdirectory()  # 选择文件夹
    if file_dir_path.strip() != '':
        dir_path.set(file_dir_path)  # 设置变量filedir1的值


# 遍历文件夹并返回所有文件列表
def file_walk(path):
    filelist = []
    for roots, dirs, files in os.walk(path, topdown=False):
        for name in files:  # 当前文件夹下所有文件（包括子文件夹下）
            filepath = os.path.join(roots, name)
            filelist.append(filepath)
    return filelist


# 选择要查看的对比结果文件
def html_check():
    html_path = askopenfilename(title=u'选择文件')

    root1 = Tk()
    root1.title("html_check")

    frame1 = Frame(root1)
    frame1.pack()
    Label(frame1, text='选择查看结果').grid(row=1, column=0, padx=100, pady=5)
    entry1 = Entry(frame1, width=50)
    entry1.grid(row=1, column=200, padx=100, pady=5)
    Button(frame1, text='浏览').grid(row=1, column=300, padx=100, pady=5)

    if not html_path:
        html_path = default
    frame2 = th3.HtmlFrame(root1, horizontal_scrollbar="auto")
    frame2.pack(fill=BOTH, expand=True)
    frame2.load_file(html_path)

    root1.mainloop()


#   输入程序所属项目
def project():
    root2 = Tk()
    root2.title("project")

    #   获得项目名，并修改菜单条
    def name():
        global project_name
        if entry1.get():
            project_name = str(entry1.get())
        root2.destroy()
        menubar.entryconfig(1, label='项目:'+project_name, command=project)

    Label(root2, text='输入项目名').grid(row=1, column=1, padx=5, pady=5)
    entry1 = Entry(root2, width=20)
    entry1.grid(row=2, column=1, padx=5, pady=5)
    Button(root2, text='完成', command=name).grid(row=3, column=1, padx=5, pady=5)

    root2.mainloop()


#   工作程序的类型
def program():
    root2 = Tk()
    root2.title("program")

    #   获得项目名，并修改菜单条
    def name():
        global program_name
        if entry1.get():
            program_name = str(entry1.get())
        root2.destroy()
        menubar.entryconfig(2, label='类型:' + program_name, command=program)

    Label(root2, text='输入程序类型').grid(row=1, column=1, padx=5, pady=5)
    entry1 = Entry(root2, width=20)
    entry1.grid(row=2, column=1, padx=5, pady=5)
    Button(root2, text='完成', command=name).grid(row=3, column=1, padx=5, pady=5)

    root2.mainloop()


# 开线程，以多线程方法运行compare程序，解决tkinter界面卡死问题
def thread_it(func, *args):
    # 将函数打包进线程
    # 创建
    t = threading.Thread(target=func, args=args)
    # 守护 !!!
    t.setDaemon(True)
    # 启动
    t.start()
    # 阻塞--卡死界面！
    # t.join()


#  为创建结果过程中添加进度条显示
def show(a, b):
    # 进度条显示
    progressbar = tkinter.ttk.Progressbar(root)
    progressbar.grid(row=6, column=1, padx=5, pady=5)
    # 进度值最大值
    progressbar['maximum'] = 100
    # 进度值初始值
    progressbar['value'] = 0
    progressbar['value'] = a / b * 100
    # 更新画面
    root.update()


# 新建对比结果的html文件
def new_html(new_folder_path, original, modified, html_name, ex):
    if os.path.exists(new_folder_path) == 0:  # 在当前路径下新建文件夹
        os.makedirs(new_folder_path)
    new_folder_path1 = os.path.join(new_folder_path, html_name)  # 当前对比文件文件夹路径
    if os.path.exists(new_folder_path1) == 0:  # 在当前路径下新建文件夹
        os.makedirs(new_folder_path1)
    # 选择生成对比结果html文件路径
    path = os.path.join(new_folder_path1, "compare_{}{}.html".format(html_name, ex))
    difference = difflib.HtmlDiff()  # 创建一个html实例
    with open(path, "w", encoding="utf-8") as fp:
        # 返回一个 html file string对象
        html = difference.make_file(fromlines=original, tolines=modified, fromdesc="Original", todesc="Modified")
        fp.write(html)

    return path


# 离线程序中必要部分替换为模板程序内容(参数皆为列表形式），使用
def text_replace(olg, old, new, dir_path, name, extension):
    # 将原程序内容替换为模板程序内容
    j = 0
    for i in olg:
        if i == old[j]:
            olg[olg.index(i)] = new[j]
            j += 1
    new_string = ''.join(olg)
    new_filepath = os.path.join(dir_path, "{}{}{}".format(name, "_new", extension))
    with open(new_filepath, 'w') as f:  # 将filename1返回为可读的字符串作为新建文件路径
        f.writelines(new_string)
    return new_string


# 数据去敏,去除不影响计算相似度的内容
def clear(s):
    string1 = re.sub(r'\d+:', '', s)    # 去除行号
    string2 = re.sub(r'\s+', '', string1)  # 去除空格
    punctuation = """，。？｡＂#＄％＆＇＊*＋－／；;,＜＝＞＠＼＾＿｀｛｜｝～､、〃〜"""
    re_punctuation = "[{}]+".format(punctuation)    # 去除标点符号
    string3 = re.sub(re_punctuation, '', string2)
    a = string3.upper()
    return a


# 抓件子程序
def pick_part(program1):
    for line in program1:
        if 'CALL JOB' in line:
            print('ok')


# 选择操作模式
def choose():
    flag = askquestion(title="操作模式", message="是否选择单个操作")
    global signal
    if flag == "yes":
        signal = 1
        mode_name.set("单个操作")
    elif flag == "no":
        signal = 2
        mode_name.set("批量操作")
    return signal


# 进行文件比对
def compare():
    current_dir = os.getcwd()
    if not save_path.get():
        # 存放所有对比文件文件夹默认路径
        default_dir = os.path.join(current_dir, 'results')
        save_path.set(default_dir)
        if os.path.exists(save_path.get()) == 0:  # 在当前路径下新建文件夹
            os.makedirs(save_path.get())
    #   菜单条查看的文件默认路径
    global default

    if signal == 1:
        if not filename1.get() or not filename2.get():
            showinfo('温馨提示', '请选择正确的操作模式！')
        else:
            with open(filename1.get(), 'r') as f1:  # 将filename1返回为可读的字符串作为新建文件路径
                a = f1.readlines()
            with open(filename2.get(), 'r') as f2:
                b = f2.readlines()

            # 存放每个程序生成结果的子文件夹名
            html_name, extension = os.path.splitext(os.path.basename(filename2.get()))
            # 生成完整对比文件并返回其存放的路径（即上述子文件夹）
            file_path = new_html(save_path.get(), a, b, html_name, '')
            default = file_path
            # 该程序的逻辑段，抬头备注和特殊语句的匹配标志！！！
            keys = keyword(extension)

            # 返回程序对比差异结果并通过文本框输出0
            retext.delete(1.0, "end")
            result = html_name + '对比结果：'
            retext.insert('end', result.center(73, '*') + '\n')

            template = ProgramTemplate(a, keys[0], keys[1], keys[2])
            prog = ProgramTemplate(b, keys[0], keys[1], keys[2])
            # 对于程序逻辑段进行比对，并生成对比文件
            x1 = template.logic_segment()
            y1 = prog.logic_segment()
            x11 = clear(''.join(x1))
            y11 = clear(''.join(y1))
            new_html(save_path.get(), x1, y1, html_name, '_mn')
            rate11 = difflib.SequenceMatcher(None, x11, y11).ratio()  # ratio = 2.0*M / T
            rate1 = '{:.3%}'.format(rate11)
            retext.insert("end", '逻辑段对比结果：{}\n'.format(rate1))

            # 对于程序抬头备注进行比对，并生成对比文件
            x2 = template.remarks()
            y2 = prog.remarks()
            new_html(save_path.get(), x2, y2, html_name, '_re')

            # 对于程序备注进行比对，并生成对比文件
            x3 = template.ex_part()
            y3 = prog.ex_part()
            new_html(save_path.get(), x3, y3, html_name, '_ex')
            x31 = clear(''.join(x3))
            y31 = clear(''.join(y3))
            rate31 = difflib.SequenceMatcher(None, x31, y31).ratio()
            rate3 = '{:.3%}'.format(rate31)
            retext.insert("end", '语句备注对比结果：{}\n'.format(rate3))

            msg1 = askyesno(title="是或否", message="是否查看对比结果")
            if msg1:
                webbrowser.open(file_path, new=1)

    elif signal == 2:
        if not filedir1.get() or not filedir2.get():
            showinfo('温馨提示', '请选择正确的操作模式！')
        else:
            # 创建表格，存放对比结果
            excel_name = '{}_results.xlsx'.format(os.path.basename(filedir2.get()))
            excel_path = os.path.join(save_path.get(), excel_name)
            workbook = Workbook()
            sheet = workbook.active
            # 表格第一行添加表头元素
            sheet.append(['工位', '程序名', '存放路径', '逻辑合格率', '备注合格率'])
            # 设置第一行元素为居中和黑色字体
            for cell in sheet[1]:
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(color='000000', bold=True)

            c = file_walk(filedir1.get())
            d = file_walk(filedir2.get())
            i = 0
            num = 2
            retext.delete(1.0, "end")
            for elem1 in c:
                i += 1
                show(i, len(c))
                for elem2 in d:
                    # 文件名匹配
                    pattern1 = re.compile(os.path.basename(elem2), re.IGNORECASE)
                    # 文件所在文件夹名匹配
                    pattern2 = re.compile(os.path.basename(os.path.dirname(elem2)), re.IGNORECASE)
                    # 选择的文件夹名匹配
                    pattern3 = re.compile(os.path.basename(filedir2.get()), re.IGNORECASE)
                    # 匹配两个文件夹内所有文件名相同且所在子文件夹相同或者所选择的文件夹不同但文件名相同
                    if ((pattern1.match(os.path.basename(elem1)) and pattern2.search(elem1)) or
                            (pattern1.match(os.path.basename(elem1)) and not pattern3.search(elem1))):
                        with open(elem1, 'r') as f3:
                            e = f3.readlines()
                        with open(elem2, 'r') as f4:
                            f = f4.readlines()

                        dir_name = os.path.basename(os.path.dirname(elem2))
                        dir_path = os.path.join(save_path.get(), dir_name)

                        # 存放每个程序生成结果的子文件夹名
                        html_name, extension = os.path.splitext(os.path.basename(elem2))
                        # 生成完整程序的对比文件并返回其存放的路径（即上述子文件夹）
                        file_path = new_html(dir_path, e, f, html_name, '')
                        result = html_name + '对比结果：'
                        retext.insert('end', result.center(73, '*') + '\n')
                        default = file_path
                        # 该程序的逻辑段，抬头备注和特殊语句的匹配标志！！！
                        keys = keyword(extension)

                        template = ProgramTemplate(e, keys[0], keys[1], keys[2])
                        prog = ProgramTemplate(f, keys[0], keys[1], keys[2])
                        # 对于程序逻辑段进行比对，并生成对比文件
                        x1 = template.logic_segment()
                        y1 = prog.logic_segment()
                        new_html(dir_path, x1, y1, html_name, '_mn')
                        # 对筛选出的语句进行数据去敏并计算符合率
                        x11 = clear(''.join(x1))
                        y11 = clear(''.join(y1))
                        rate11 = difflib.SequenceMatcher(None, x11, y11).ratio()
                        rate1 = '{:.3%}'.format(rate11)
                        retext.insert("end", '逻辑段对比结果：{}\n'.format(rate1))

                        # 对于程序抬头备注进行比对，并生成对比文件
                        x2 = template.remarks()
                        y2 = prog.remarks()
                        new_html(dir_path, x2, y2, html_name, '_re')

                        # 对于程序特殊备注进行比对，并生成对比文件
                        x3 = template.ex_part()
                        y3 = prog.ex_part()
                        new_html(dir_path, x3, y3, html_name, '_ex')
                        # 对筛选出的语句进行数据去敏并计算符合率
                        x31 = clear(''.join(x3))
                        y31 = clear(''.join(y3))
                        rate31 = difflib.SequenceMatcher(None, x31, y31).ratio()
                        rate3 = '{:.3%}'.format(rate31)
                        retext.insert("end", '语句备注对比结果：{}\n\n'.format(rate3))

                        # 表格存放具体的生成结果
                        result = [dir_name, html_name, file_path, rate1, rate3]
                        for j in range(1, 6):
                            sheet.cell(row=num, column=j, value=result[j-1])
                        num += 1
                        sheet.title = 'sheet1'

            showinfo('完成提示', '结果已生成！')
            # excel文件保存
            # 设置其余元素背景颜色为黄色
            # yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for row in sheet.iter_rows(min_row=2, min_col=4, max_col=5):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
            workbook.save(excel_path)

    else:
        showinfo('温馨提示', '请选择操作模式！')


# 创建菜单栏
menubar = Menu(root)
# 将菜单栏添加到窗口
root.config(menu=menubar)
# 为菜单条添加具体功能
menubar.add_command(label=project_name, command=project)
menubar.add_command(label=program_name, command=program)
menubar.add_command(label='查看', command=html_check)

# 选择原文件
Label(root, text='选择模版文件').grid(row=1, column=0, padx=5, pady=5)
Entry(root, textvariable=filename1, width=40).grid(row=1, column=1, padx=5, pady=5)
Button(root, text='浏览', command=lambda: input_file(filename1)).grid(row=1, column=2, padx=5, pady=5)
# 选择对比文件
Label(root, text='选择对比文件').grid(row=2, column=0, padx=5, pady=5)
Entry(root, textvariable=filename2, width=40).grid(row=2, column=1, padx=5, pady=5)
Button(root, text='浏览', command=lambda: input_file(filename2)).grid(row=2, column=2, padx=5, pady=5)
# 选择原文件所在文件夹
Label(root, text='选择模版文件夹').grid(row=3, column=0, padx=5, pady=5)
Entry(root, textvariable=filedir1, width=40).grid(row=3, column=1, padx=5, pady=5)
Button(root, text='浏览', command=lambda: input_dir(filedir1)).grid(row=3, column=2, padx=5, pady=5)
# 选择对比文件所在文件夹
Label(root, text='选择对比文件夹').grid(row=4, column=0, padx=5, pady=5)
Entry(root, textvariable=filedir2, width=40).grid(row=4, column=1, padx=5, pady=5)
Button(root, text='浏览', command=lambda: input_dir(filedir2)).grid(row=4, column=2, padx=5, pady=5)
# 选择对比结果存放的文件夹路径
Label(root, text='选择存放位置').grid(row=5, column=0, padx=5, pady=5)
Entry(root, textvariable=save_path, width=40).grid(row=5, column=1, padx=5, pady=5)
Button(root, text='浏览', command=lambda: input_dir(save_path)).grid(row=5, column=2, padx=5, pady=5)
# 选择操作模式
Button(root, textvariable=mode_name, command=choose).grid(row=6, column=0, padx=5, pady=5)
# 生成对比结果
Button(root, text='生成对比结果', command=lambda: thread_it(compare)).grid(row=6, column=2, padx=5, pady=5)
# 文本框显示结果
retext = Text(root)
retext.grid(row=7, column=0, rowspan=3, columnspan=3, padx=5, pady=5)
# 进入消息循环
root.mainloop()
